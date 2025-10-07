from __future__ import annotations
from typing import Dict
import unicodedata, re, datetime
from openpyxl.worksheet.worksheet import Worksheet

def _strip_accents(s: str) -> str:
    if not isinstance(s, str): 
        return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))

def _norm_header(s: str) -> str:
    s = _strip_accents(s or "").lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s

def find_header_row_and_map(ws: Worksheet, max_scan_rows: int = 10):
    """Detecta linha de cabeçalho e retorna (header_row_idx, {'data': c, 'modalidade': c, 'materia': c})."""
    for r in range(1, max_scan_rows+1):
        row_vals = [_norm_header((ws.cell(r, c).value or "")) for c in range(1, ws.max_column+1)]
        cand = {"data": None, "modalidade": None, "materia": None}
        for c, hv in enumerate(row_vals, start=1):
            if hv in ("data", "dia"):
                cand["data"] = c
            elif hv.startswith("modalidade"):
                cand["modalidade"] = c
            elif hv in ("materia lecionada", "materia", "conteudo", "conteudo lecionado", "descricao", "descrição"):
                cand["materia"] = c
        if all(cand.values()):
            return r, cand
    return 1, {"data": 1, "modalidade": 2, "materia": 4}

def fmt_date_ddmmyyyy(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    s = re.sub(r"[^\d/]", "", s)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        d, mth, y = m.groups()
        y = y.zfill(4) if len(y) == 4 else ("20" + y.zfill(2))
        return f"{d.zfill(2)}/{mth.zfill(2)}/{y}"
    return None

def mod_to_suffix(mod: str) -> str | None:
    m = _strip_accents((mod or "")).lower()
    if "teor" in m:  # teórica
        return "T"
    if "prat" in m:  # prática
        return "P"
    return None

def process_worksheet(ws: Worksheet, validate_value_map) -> tuple[Dict[str, str], dict]:
    """Lê uma worksheet e devolve (norm_map, stats)."""
    header_row, cols = find_header_row_and_map(ws)
    c_data, c_mod, c_mat = cols["data"], cols["modalidade"], cols["materia"]

    new_items: Dict[str, str] = {}
    imported = 0
    skipped = 0
    overwritten = 0

    for r in range(header_row + 1, ws.max_row + 1):
        v_date = ws.cell(r, c_data).value
        v_mod = ws.cell(r, c_mod).value
        v_text = ws.cell(r, c_mat).value

        key_date = fmt_date_ddmmyyyy(v_date)
        suf = mod_to_suffix(v_mod)
        text = (str(v_text).strip() if v_text is not None else "")

        if not key_date or not suf or not text:
            skipped += 1
            continue

        key = f"{key_date} -{suf}"
        if key in new_items:
            overwritten += 1
        new_items[key] = text
        imported += 1

    norm, errors = validate_value_map(new_items)
    stats = {
        "imported": imported,
        "skipped": skipped,
        "overwritten_in_lot": overwritten,
        "errors": errors,
        "valid": len(norm),
    }
    return norm, stats