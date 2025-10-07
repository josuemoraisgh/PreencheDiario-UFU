# ui.py
from __future__ import annotations
import json
import threading
from typing import Dict, Optional
from tkinter import (
    Tk, Frame, Button, Listbox, Text, Scrollbar, END, SINGLE, BOTH, LEFT, RIGHT,
    Y, X, TOP, BOTTOM, filedialog, simpledialog, messagebox, Toplevel, Label, Entry, StringVar
)
from tkinter import ttk
from openpyxl import load_workbook
import unicodedata, datetime, re, calendar


from selenium.webdriver.remote.webdriver import WebDriver

# imports do seu projeto
from services.drivers import create_driver
from services.utils import GET_URL, validate_value_map, preview_text
from services.diario import get_current_turma_info, fill_entries, try_click_save
from services.cookies import read_cookie_file, test_cookie_header

# --- Helper drop-in: use Centerlevel(self) no lugar de Toplevel(self) ---
def Centerlevel(parent, *, transient=True, grab=True, center_on_parent=True, topmost_pulse=True, **kwargs):
    """
    Cria um Toplevel e centraliza automaticamente quando aparecer.
    Uso: win = Centerlevel(self)  # substitui Toplevel(self)
    Parâmetros:
      - transient: chama win.transient(parent) (fica acima do pai)
      - grab:      chama win.grab_set() (modal)
      - center_on_parent: True = centra na janela pai; False = centro da tela
      - topmost_pulse: traz à frente rapidamente para evitar ficar atrás de outras janelas
    """
    win = Toplevel(parent, **kwargs)

    if transient:
        try: win.transient(parent)
        except Exception: pass

    # Não mostrar até posicionar
    try: win.withdraw()
    except Exception: pass

    # Centralização (usa PlaceWindow se disponível; senão, calcula manual)
    def _center_once():
        try:
            win.update_idletasks()
            try:
                # Tk 8.6+: centraliza de forma robusta
                win.eval(f'tk::PlaceWindow {win.winfo_pathname(win.winfo_id())} center')
                if center_on_parent:
                    # PlaceWindow centra na tela; se quiser centrar no pai, faz manual:
                    raise Exception("force-manual-parent-center")
            except Exception:
                # Fallback manual: centro no pai ou na tela
                if center_on_parent and parent is not None:
                    parent.update_idletasks()
                    px, py = parent.winfo_rootx(), parent.winfo_rooty()
                    pw, ph = parent.winfo_width(), parent.winfo_height()
                    w, h = win.winfo_reqwidth(), win.winfo_reqheight()
                    if pw <= 1 or ph <= 1:  # pai ainda não medido
                        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
                        x, y = (sw - w)//2, (sh - h)//2
                    else:
                        x, y = px + (pw - w)//2, py + (ph - h)//2
                else:
                    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
                    w, h = win.winfo_reqwidth(), win.winfo_reqheight()
                    x, y = (sw - w)//2, (sh - h)//2
                win.geometry(f"+{x}+{y}")

            # Mostrar após posicionar
            try: win.deiconify()
            except Exception: pass

            # Traz à frente rapidamente
            if topmost_pulse:
                try:
                    win.lift()
                    win.attributes("-topmost", True)
                    # solta o topmost para não “grudar” no topo
                    win.after(10, lambda: win.attributes("-topmost", False))
                except Exception:
                    pass

            # Modal
            if grab:
                try: win.grab_set()
                except Exception: pass

        except Exception:
            # fallback final: mostra mesmo assim
            try: win.deiconify()
            except Exception: pass

    # Centraliza quando a janela estiver pronta
    # (chama duas vezes para capturar ajustes de layout tardios)
    win.after(0, _center_once)
    win.after(120, _center_once)

    return win


class App(Tk):
    def __init__(self):
        super().__init__()
        self.title("UFU Diário – Preenchimento (visual)")
        self.geometry("1200x720")
        self.minsize(1000, 600)

        # Estado
        self.driver: Optional[WebDriver] = None
        self.value_map: Dict[str, str] = {}
        self.current_path: Optional[str] = None

        # Browser selecionado
        self.browser_var = StringVar(value="edge")

        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        # Top bar
        top = Frame(self); top.pack(side=TOP, fill=X)

        # Browser selector
        ttk.Label(top, text="Browser:").pack(side=LEFT, padx=(6, 2), pady=6)
        self.cbo_browser = ttk.Combobox(
            top, textvariable=self.browser_var, state="readonly",
            values=["edge", "chrome", "firefox"], width=10
        )
        self.cbo_browser.pack(side=LEFT, padx=(0, 8), pady=6)

        self.btn_open_browser = Button(top, text="Start Navegador", command=self.on_open_browser)
        self.btn_open_browser.pack(side=LEFT, padx=4, pady=6)

        self.btn_load_json = Button(top, text="Carregar Dados", command=self.on_load_json)
        self.btn_load_json.pack(side=LEFT, padx=4, pady=6)

        self.btn_import_excel = Button(top, text="Importar Excel", command=self.on_import_excel)
        self.btn_import_excel.pack(side=LEFT, padx=4, pady=6)


        # Main split
        main = Frame(self); main.pack(side=TOP, fill=BOTH, expand=True)
        left = Frame(main, width=520); left.pack(side=LEFT, fill=BOTH, expand=True)
        right = Frame(main); right.pack(side=RIGHT, fill=BOTH, expand=True)

        # Left panel (lista)
        self.listbox = Listbox(left, selectmode=SINGLE)
        self.listbox.pack(side=TOP, fill=BOTH, expand=True, padx=6, pady=6)

        left_btns = Frame(left); left_btns.pack(side=BOTTOM, fill=X, padx=6, pady=6)

        self.btn_add = Button(left_btns, text="Insert", command=self.on_add_item)
        self.btn_add.pack(side=LEFT, padx=4)

        self.btn_edit = Button(left_btns, text="Edit", command=self.on_edit_item)
        self.btn_edit.pack(side=LEFT, padx=4)

        self.btn_remove = Button(left_btns, text="Remove", command=self.on_remove_item)
        self.btn_remove.pack(side=LEFT, padx=4)

        self.btn_shift = Button(left_btns, text="Ajustar datas (±)", command=self.on_shift_dates)
        self.btn_shift.pack(side=LEFT, padx=4)

        self.btn_save_json = Button(left_btns, text="Salvar JSON", command=self.on_save_json)
        self.btn_save_json.pack(side=LEFT, padx=4)

        self.btn_fill = Button(left_btns, text="Preencher diário", command=self.on_fill, state="disabled")
        self.btn_fill.pack(side=RIGHT, padx=4)

        # Right panel (Logs)
        self.logs = Text(right, wrap="word", state="disabled")
        sb = Scrollbar(right, command=self.logs.yview)
        self.logs.configure(yscrollcommand=sb.set)
        self.logs.pack(side=LEFT, fill=BOTH, expand=True, padx=6, pady=6)
        sb.pack(side=RIGHT, fill=Y)

        self._log("Pronto. Abra o navegador, carregue dados.json e navegue até a turma.")

    # ---------- Helpers ----------
    def _log(self, msg: str):
        self.logs.configure(state="normal")
        self.logs.insert(END, msg + "\n")
        self.logs.see(END)
        self.logs.configure(state="disabled")

    def _log_clear(self):
        self.logs.configure(state="normal")
        self.logs.delete("1.0", END)
        self.logs.configure(state="disabled")

    def _refresh_listbox(self):
        self.listbox.delete(0, END)
        for k, v in self.value_map.items():
            self.listbox.insert(END, f"{k}: {preview_text(v)}")

    def _validate_ready(self):
        ready = (self.driver is not None) and bool(self.value_map)
        self.btn_fill.configure(state=("normal" if ready else "disabled"))

    # ---------- Top bar actions ----------
    def on_open_browser(self):
        def _run():
            try:
                browser = (self.browser_var.get() or "edge").strip().lower()
                if self.driver is None:
                    self._log(f"[UI] Abrindo {browser.title()}...")
                    self.driver = create_driver(browser=browser, logger=self._log)
                self._log(f"[UI] Navegando para: {GET_URL}")
                self.driver.get(GET_URL)
                self._validate_ready()
            except Exception as e:
                self._log(f"[ERRO] Falha ao abrir navegador: {e}")
        threading.Thread(target=_run, daemon=True).start()

    def on_load_json(self):
        path = filedialog.askopenfilename(title="Escolha dados.json", filetypes=[("JSON", "*.json")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            norm, errors = validate_value_map(raw)
            self._log_clear()
            if errors:
                self._log("⚠ Erros ao validar dados.json:")
                for k, e in errors.items():
                    self._log(f" - {k}: {e}")
            self.value_map = norm
            self.current_path = path
            self._refresh_listbox()
            self._log("✔ dados.json carregado. Itens:")
            for k, v in self.value_map.items():
                self._log(f" - {k}: {preview_text(v)}")
            self._validate_ready()
        except Exception as e:
            self._log_clear()
            self._log(f"[ERRO] Falha ao carregar JSON: {e}")

    # ---------- CRUD dos itens ----------
    def _edit_dialog(self, initial_key: str, initial_text: str) -> Optional[tuple[str, str]]:
        """Modal simples para editar chave e texto; retorna (new_key, new_text) ou None."""
        win = Centerlevel(self)
        win.title("Editar item")
        win.transient(self)
        win.grab_set()

        Label(win, text="Chave (DD/MM/AAAA -X):").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        var_key = StringVar(value=initial_key)
        ent_key = Entry(win, textvariable=var_key, width=40)
        ent_key.grid(row=1, column=0, sticky="we", padx=8)

        Label(win, text="Texto:").grid(row=2, column=0, sticky="w", padx=8, pady=(8, 4))
        # Para simplicidade, usar Entry também (mantém comportamento do original).
        # Se quiser multiline, troque para Text com scroll.
        var_txt = StringVar(value=initial_text)
        ent_txt = Entry(win, textvariable=var_txt, width=80)
        ent_txt.grid(row=3, column=0, sticky="we", padx=8)

        btns = Frame(win); btns.grid(row=4, column=0, sticky="e", padx=8, pady=8)
        result: list[Optional[tuple[str, str]]] = [None]

        def _ok():
            result[0] = (var_key.get().strip(), var_txt.get())
            win.destroy()

        def _cancel():
            win.destroy()

        Button(btns, text="Cancelar", command=_cancel).pack(side=RIGHT, padx=4)
        Button(btns, text="OK", command=_ok).pack(side=RIGHT, padx=4)

        win.columnconfigure(0, weight=1)
        ent_key.focus_set()
        self.wait_window(win)
        return result[0]

    def on_edit_item(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showinfo("Editar", "Selecione um item na lista.")
            return
        idx = sel[0]
        old_key = list(self.value_map.keys())[idx]
        old_text = self.value_map[old_key]

        res = self._edit_dialog(old_key, old_text)
        if not res:
            return
        new_key, new_text = res

        # Validação: normaliza e checa conflitos
        norm, errors = validate_value_map({new_key: new_text})
        if errors:
            msg = "\n".join(f"{k}: {e}" for k, e in errors.items())
            messagebox.showerror("Erro de validação", msg)
            return

        new_key_norm = list(norm.keys())[0]
        new_text_norm = norm[new_key_norm]

        # Se a chave mudou, remover a antiga
        if new_key_norm != old_key:
            # Conflito?
            if new_key_norm in self.value_map and new_key_norm != old_key:
                messagebox.showerror("Conflito", f"A chave {new_key_norm!r} já existe.")
                return
            del self.value_map[old_key]

        self.value_map[new_key_norm] = new_text_norm
        # Recria a lista mantendo ordem por chave (opcional)
        self.value_map = dict(sorted(self.value_map.items(), key=lambda kv: kv[0]))
        self._refresh_listbox()
        self._log(f"[UI] Item editado: {new_key_norm}")
        self._validate_ready()

    def on_add_item(self):
        # Pergunta chave e texto (mantendo padrão simples)
        key = simpledialog.askstring("Nova entrada", "Informe a chave (DD/MM/AAAA -X):", parent=self)
        if not key:
            return
        val = simpledialog.askstring("Texto", "Informe o texto:", parent=self) or ""
        norm, errors = validate_value_map({key: val})
        if errors:
            msg = "\n".join(f"{k}: {e}" for k, e in errors.items())
            messagebox.showerror("Erro de validação", msg)
            return

        nk = list(norm.keys())[0]
        if nk in self.value_map:
            messagebox.showerror("Conflito", f"A chave {nk!r} já existe.")
            return

        self.value_map[nk] = norm[nk]
        self.value_map = dict(sorted(self.value_map.items(), key=lambda kv: kv[0]))
        self._refresh_listbox()
        self._log(f"[UI] Item adicionado: {nk}")
        self._validate_ready()

    def on_remove_item(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showinfo("Remover", "Selecione um item na lista.")
            return
        idx = sel[0]
        key = list(self.value_map.keys())[idx]
        if not messagebox.askyesno("Confirmar remoção", f"Remover a entrada '{key}'?"):
            return
        del self.value_map[key]
        self._refresh_listbox()
        self._log(f"[UI] Item removido: {key}")
        self._validate_ready()

    def on_save_json(self):
        path = filedialog.asksaveasfilename(
            title="Salvar JSON", defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile=(self.current_path or "dados.json")
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.value_map, f, ensure_ascii=False, indent=2)
            self._log(f"[UI] Salvo em: {path}")
        except Exception as e:
            self._log(f"[ERRO] Falha ao salvar JSON: {e}")

    # ---------- Preenchimento ----------
    def on_fill(self):
        if not self.driver:
            messagebox.showerror("Navegador", "Abra o navegador primeiro.")
            return
        if not self.value_map:
            messagebox.showerror("Dados", "Carregue um dados.json válido.")
            return

        n = len(self.value_map)
        resp = messagebox.askyesno(
            "Confirmar",
            f"Você vai preencher {n} itens na turma atual (página aberta no navegador). Continuar?"
        )
        if not resp:
            return

        def _run():
            try:
                info = get_current_turma_info(self.driver)
                if not info:
                    self._log("⚠ Não consegui detectar idTurma/tipo na página atual. Abra a página do diário e tente novamente.")
                    return
                self._log("Iniciando preenchimento visual...")
                ok, fail = fill_entries(self.driver, self.value_map, self._log)
                try_click_save(self.driver, self._log)
                self._log(f"Preenchimento concluído: {ok} ok, {fail} não encontrado.")
            except Exception as e:
                self._log(f"[ERRO] Falha no preenchimento: {e}")

        threading.Thread(target=_run, daemon=True).start()
        
    # ---------- Importação Excel ----------
    def _strip_accents(self, s: str) -> str:
        if not isinstance(s, str): 
            return ""
        return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))

    def _norm_header(self, s: str) -> str:
        s = self._strip_accents(s or "").lower().strip()
        s = re.sub(r"\s+", " ", s)
        return s

    def _choose_from_list(self, title: str, options: list[str]) -> Optional[str]:
        """Modal simples com Combobox para escolher um item de options."""
        if not options:
            return None
        win = Centerlevel(self)
        win.title(title)
        win.transient(self); win.grab_set()

        Label(win, text=title).grid(row=0, column=0, padx=8, pady=(8,4), sticky="w")
        var = StringVar(value=options[0])
        cb = ttk.Combobox(win, textvariable=var, values=options, state="readonly", width=40)
        cb.grid(row=1, column=0, padx=8, pady=4, sticky="we")

        res = [None]
        def _ok(): res[0] = var.get(); win.destroy()
        def _cancel(): win.destroy()
        btns = Frame(win); btns.grid(row=2, column=0, sticky="e", padx=8, pady=8)
        Button(btns, text="Cancelar", command=_cancel).pack(side=RIGHT, padx=4)
        Button(btns, text="OK", command=_ok).pack(side=RIGHT, padx=4)

        win.columnconfigure(0, weight=1)
        cb.focus_set()
        self.wait_window(win)
        return res[0]

    def _find_header_row_and_map(self, ws, max_scan_rows: int = 10):
        """Detecta linha de cabeçalho e retorna (header_row_idx, map_dict).
        map_dict = {'data': col_idx, 'modalidade': col_idx, 'materia': col_idx}
        """
        wants = {"data": None, "modalidade": None, "materia": None}
        for r in range(1, max_scan_rows+1):
            row_vals = [self._norm_header((ws.cell(r, c).value or "")) for c in range(1, ws.max_column+1)]
            # procure colunas
            cand = {"data": None, "modalidade": None, "materia": None}
            for c, hv in enumerate(row_vals, start=1):
                if hv in ("data", "dia"):
                    cand["data"] = c
                elif hv.startswith("modalidade"):
                    cand["modalidade"] = c
                elif hv in ("materia lecionada", "materia", "conteudo", "conteudo lecionado", "descricao", "descrição"):
                    cand["materia"] = c
            if all(cand.values()):
                return r, cand
        # fallback: primeira linha
        return 1, {"data": 1, "modalidade": 2, "materia": 4}

    def _fmt_date_ddmmyyyy(self, val) -> Optional[str]:
        """Aceita datetime/date ou string; retorna DD/MM/AAAA."""
        if val is None or val == "":
            return None
        if isinstance(val, (datetime.date, datetime.datetime)):
            return val.strftime("%d/%m/%Y")
        s = str(val).strip()
        s = re.sub(r"[^\d/]", "", s)  # mantém apenas dígitos e /
        # tenta DD/MM/AAAA
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
        if m:
            d, mth, y = m.groups()
            y = y.zfill(4) if len(y) == 4 else ("20" + y.zfill(2))
            return f"{d.zfill(2)}/{mth.zfill(2)}/{y}"
        return None

    def _mod_to_suffix(self, mod: str) -> Optional[str]:
        m = self._strip_accents((mod or "")).lower()
        if "teor" in m:  # teórica
            return "T"
        if "prat" in m:  # prática
            return "P"
        return None

    def on_import_excel(self):
        path = filedialog.askopenfilename(
            title="Selecione a planilha",
            filetypes=[("Excel", "*.xlsx;*.xlsm;*.xltx;*.xltm"), ("Todos", "*.*")]
        )
        if not path:
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            messagebox.showerror("Excel", f"Não consegui abrir o arquivo:\n{e}")
            return

        # escolher aba
        sheetnames = list(wb.sheetnames)
        sheet = sheetnames[0] if len(sheetnames) == 1 else self._choose_from_list("Escolha a aba", sheetnames)
        if not sheet:
            return
        ws = wb[sheet]

        # detectar cabeçalhos
        header_row, cols = self._find_header_row_and_map(ws)
        c_data, c_mod, c_mat = cols["data"], cols["modalidade"], cols["materia"]

        # varre linhas
        new_items = {}
        imported = 0
        skipped = 0
        overwritten = 0

        for r in range(header_row + 1, ws.max_row + 1):
            v_date = ws.cell(r, c_data).value
            v_mod = ws.cell(r, c_mod).value
            v_text = ws.cell(r, c_mat).value

            key_date = self._fmt_date_ddmmyyyy(v_date)
            suf = self._mod_to_suffix(v_mod)
            text = (str(v_text).strip() if v_text is not None else "")

            # regras de ignorar
            if not key_date or not suf or not text:
                skipped += 1
                continue

            key = f"{key_date} -{suf}"
            if key in new_items:
                overwritten += 1  # dentro do próprio lote
            new_items[key] = text
            imported += 1

        wb.close()

        # valida e normaliza com as regras já do app
        norm, errors = validate_value_map(new_items)
        if errors:
            self._log("⚠ Erros ao validar itens importados:")
            for k, e in errors.items():
                self._log(f" - {k}: {e}")

        # mescla com o que já tem (sobrescreve duplicadas)
        dup_on_merge = sum(1 for k in norm if k in self.value_map)
        self.value_map.update(norm)

        # ordena por chave (opcional, ajuda na visualização)
        self.value_map = dict(sorted(self.value_map.items(), key=lambda kv: kv[0]))

        # UI
        self._refresh_listbox()
        self._validate_ready()

        # logs
        self._log("✔ Importação Excel concluída.")
        self._log(f"   Arquivo: {path}")
        self._log(f"   Aba: {sheet}")
        self._log(f"   Itens válidos: {len(norm)} | Linhas ignoradas: {skipped} | Sobrescritas neste lote: {overwritten} | Sobrescritas na mesclagem: {dup_on_merge}")
        self._log("   Preview:")
        preview_count = 0
        for k in list(norm.keys()):
            self._log(f"   - {k}: {preview_text(norm[k])}")
            preview_count += 1
            if preview_count >= 5:
                break

    # ---------- Ajuste em massa das datas das chaves ----------
    def _parse_key_date_suffix(self, key: str):
        """Extrai (date, suffix) de 'DD/MM/AAAA -X...'. Retorna (datetime.date, 'X...') ou (None, None) se inválido."""
        m = re.match(r"^\s*(\d{2})/(\d{2})/(\d{4})\s*-\s*(.+)\s*$", key or "")
        if not m:
            return None, None
        d, mth, y, suf = m.groups()
        try:
            dt = datetime.date(int(y), int(mth), int(d))
            return dt, suf.strip()
        except Exception:
            return None, None

    def _last_day_of_month(self, year: int, month: int) -> int:
        return calendar.monthrange(year, month)[1]

    def _add_months(self, dt: datetime.date, n: int) -> datetime.date:
        """Soma n meses (pode ser negativo). Ajusta para último dia do mês quando necessário."""
        y = dt.year + (dt.month - 1 + n) // 12
        m = (dt.month - 1 + n) % 12 + 1
        d = min(dt.day, self._last_day_of_month(y, m))
        return datetime.date(y, m, d)

    def _add_years(self, dt: datetime.date, n: int) -> datetime.date:
        """Soma n anos (pode ser negativo). Ajusta 29/fev para último dia do mês quando necessário."""
        y = dt.year + n
        m = dt.month
        d = min(dt.day, self._last_day_of_month(y, m))
        return datetime.date(y, m, d)

    def _format_key(self, dt: datetime.date, suffix: str) -> str:
        return f"{dt.strftime('%d/%m/%Y')} -{suffix}"

    def on_shift_dates(self):
        """Abre modal para escolher unidade e valor; aplica em todas as chaves."""
        win = Centerlevel(self)
        win.title("Ajustar datas das chaves")
        win.transient(self)
        win.grab_set()

        # Unidade
        Label(win, text="Unidade:").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        unit_var = StringVar(value="Dias")
        cbo = ttk.Combobox(win, textvariable=unit_var, values=["Dias", "Meses", "Anos"], state="readonly", width=12)
        cbo.grid(row=0, column=1, sticky="w", padx=8, pady=(8, 4))

        # Valor
        Label(win, text="Valor (inteiro, pode ser negativo):").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        val_var = StringVar(value="1")
        ent = Entry(win, textvariable=val_var, width=12)
        ent.grid(row=1, column=1, sticky="w", padx=8, pady=4)

        # Botões
        btns = Frame(win)
        btns.grid(row=2, column=0, columnspan=2, sticky="e", padx=8, pady=8)

        def _apply():
            # valida entrada
            try:
                amount = int(val_var.get().strip())
            except ValueError:
                messagebox.showerror("Valor inválido", "Informe um inteiro (ex.: -7, 0, 15).")
                return
            unit = unit_var.get()

            if not self.value_map:
                messagebox.showinfo("Nada a ajustar", "Não há itens na lista.")
                win.destroy()
                return

            # aplica transformações
            new_map = {}
            invalid = 0
            changed = 0
            overwritten = 0

            for key, text in self.value_map.items():
                dt, suf = self._parse_key_date_suffix(key)
                if not dt:
                    invalid += 1
                    continue

                if unit == "Dias":
                    new_dt = dt + datetime.timedelta(days=amount)
                elif unit == "Meses":
                    new_dt = self._add_months(dt, amount)
                else:  # "Anos"
                    new_dt = self._add_years(dt, amount)

                new_key = self._format_key(new_dt, suf)
                if new_key in new_map:
                    # colisão dentro do lote
                    overwritten += 1
                new_map[new_key] = text
                changed += 1

            # mescla no dicionário principal (sobrescreve se já existir)
            # se preferir preservar os antigos, troque para: tmp = self.value_map.copy(); tmp.update(new_map); self.value_map = tmp
            dup_on_merge = sum(1 for k in new_map if k in self.value_map and self.value_map.get(k) != new_map[k])
            self.value_map.update(new_map)

            # ordena para visualização
            self.value_map = dict(sorted(self.value_map.items(), key=lambda kv: kv[0]))

            # UI & logs
            self._refresh_listbox()
            self._validate_ready()
            self._log(f"✔ Ajuste concluído: {changed} chaves alteradas | inválidas: {invalid} | sobrescritas no lote: {overwritten} | sobrescritas na mesclagem: {dup_on_merge}")
            self._log(f"   Unidade: {unit} | Valor: {amount:+d}")
            # preview
            self._log("   Preview de 5 chaves após ajuste:")
            count = 0
            for k in self.value_map.keys():
                self._log(f"   - {k}")
                count += 1
                if count >= 5:
                    break

            win.destroy()

        def _cancel():
            win.destroy()

        Button(btns, text="Cancelar", command=_cancel).pack(side=RIGHT, padx=4)
        Button(btns, text="Aplicar", command=_apply).pack(side=RIGHT, padx=4)

        win.columnconfigure(0, weight=0)
        win.columnconfigure(1, weight=1)
        cbo.focus_set()
        self.wait_window(win)

    def on_shift_dates(self):
        """Ajusta datas das chaves com filtro (Todas / só T / só P) e unidade (Dias/Meses/Anos)."""
        win = Centerlevel(self)
        win.title("Ajustar datas das chaves")
        win.transient(self)
        win.grab_set()

        # Unidade
        Label(win, text="Unidade:").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        unit_var = StringVar(value="Dias")
        cbo_unit = ttk.Combobox(win, textvariable=unit_var,
                                values=["Dias", "Meses", "Anos"],
                                state="readonly", width=12)
        cbo_unit.grid(row=0, column=1, sticky="w", padx=8, pady=(8, 4))

        # Valor
        Label(win, text="Valor (inteiro, pode ser negativo):").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        val_var = StringVar(value="1")
        ent_val = Entry(win, textvariable=val_var, width=12)
        ent_val.grid(row=1, column=1, sticky="w", padx=8, pady=4)

        # Filtro (Todas / T / P)
        Label(win, text="Aplicar em:").grid(row=2, column=0, sticky="w", padx=8, pady=4)
        filt_var = StringVar(value="Todas")
        cbo_filt = ttk.Combobox(win, textvariable=filt_var,
                                values=["Todas", "Só T (Teóricas)", "Só P (Práticas)"],
                                state="readonly", width=16)
        cbo_filt.grid(row=2, column=1, sticky="w", padx=8, pady=4)

        # Botões
        btns = Frame(win)
        btns.grid(row=3, column=0, columnspan=2, sticky="e", padx=8, pady=8)

        def _apply():
            # valida entrada
            try:
                amount = int(val_var.get().strip())
            except ValueError:
                messagebox.showerror("Valor inválido", "Informe um inteiro (ex.: -7, 0, 15).")
                return
            unit = unit_var.get()
            filt = filt_var.get()  # "Todas" | "Só T (Teóricas)" | "Só P (Práticas)"

            if not self.value_map:
                messagebox.showinfo("Nada a ajustar", "Não há itens na lista.")
                win.destroy()
                return

            # Transforma tudo criando um NOVO dicionário (evita duplicar antigas)
            result_map = {}
            changed = 0
            invalid = 0
            skipped_filter = 0
            overwritten_lote = 0

            for key, text in list(self.value_map.items()):
                dt, suf_full = self._parse_key_date_suffix(key)
                if not dt:
                    # chave fora do padrão, mantém como está
                    result_map[key] = text
                    invalid += 1
                    continue

                suf_letter = self._suffix_letter(suf_full)  # P/T/…

                # aplica filtro
                if filt.startswith("Só T") and suf_letter != "T":
                    result_map[key] = text
                    skipped_filter += 1
                    continue
                if filt.startswith("Só P") and suf_letter != "P":
                    result_map[key] = text
                    skipped_filter += 1
                    continue

                # calcula nova data
                if unit == "Dias":
                    new_dt = dt + datetime.timedelta(days=amount)
                elif unit == "Meses":
                    new_dt = self._add_months(dt, amount)
                else:  # "Anos"
                    new_dt = self._add_years(dt, amount)

                new_key = self._format_key(new_dt, suf_full)

                if new_key in result_map:
                    overwritten_lote += 1  # colisão dentro do resultado
                result_map[new_key] = text
                changed += 1

            # aplica resultado
            self.value_map = dict(sorted(result_map.items(), key=lambda kv: kv[0]))

            # UI & logs
            self._refresh_listbox()
            self._validate_ready()
            self._log(
                f"✔ Ajuste concluído: {changed} alteradas | inválidas: {invalid} | filtradas: {skipped_filter} | "
                f"sobrescritas no lote: {overwritten_lote}"
            )
            self._log(f"   Unidade: {unit} | Valor: {amount:+d} | Filtro: {filt}")
            # preview
            self._log("   Preview de 5 chaves após ajuste:")
            for i, k in enumerate(self.value_map.keys()):
                if i >= 5: break
                self._log(f"   - {k}")

            win.destroy()

        def _cancel():
            win.destroy()

        Button(btns, text="Cancelar", command=_cancel).pack(side=RIGHT, padx=4)
        Button(btns, text="Aplicar", command=_apply).pack(side=RIGHT, padx=4)

        win.columnconfigure(0, weight=0)
        win.columnconfigure(1, weight=1)
        cbo_unit.focus_set()
        self.wait_window(win)

def run():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    run()
